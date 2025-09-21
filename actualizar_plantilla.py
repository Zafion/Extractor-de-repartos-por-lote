import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

SOMME_KEY = "Somme"
CODE_LOT_LABEL = "Code Lot"

def primera_celda_no_vacia_de_fila(ws, r, col_max_posible=None):
    cmax = col_max_posible or ws.max_column
    for c in range(1, cmax + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip() != "":
            return c, str(v).strip()
    return None, ""

def encontrar_filas_limite(ws):
    """
    - fila_header: fila cuya PRIMERA celda no vacía es exactamente "Date".
    - fila_somme: primera fila debajo del header cuyo primer valor no vacío empieza por "Somme".
    - col_ultima: última columna con dato en la fila de encabezados (ancho de la tabla origen).
    """
    fila_header = None
    col_ultima = None

    # 1) Header: primera celda no vacía == "Date"
    for r in range(1, ws.max_row + 1):
        cidx, first_val = primera_celda_no_vacia_de_fila(ws, r)
        if cidx is None:
            continue
        if first_val == "Date":
            fila_header = r
            break
    if not fila_header:
        raise ValueError("No se encontró el encabezado cuya primera celda sea exactamente 'Date'.")

    # 2) Última col con dato en el header
    last_col_idx = 0
    for cell in ws[fila_header]:
        if (cell.value is not None) and (str(cell.value).strip() != ""):
            last_col_idx = max(last_col_idx, cell.column)
    if last_col_idx == 0:
        raise ValueError("El encabezado está vacío.")
    col_ultima = last_col_idx

    # 3) Fila "Somme" debajo del header
    fila_somme = None
    for r in range(fila_header + 1, ws.max_row + 1):
        cidx, first_val = primera_celda_no_vacia_de_fila(ws, r, col_ultima)
        if cidx is None:
            continue
        if first_val.startswith(SOMME_KEY):
            fila_somme = r
            break
    if not fila_somme:
        raise ValueError("No se encontró la fila que empieza por 'Somme'.")
    if fila_somme <= fila_header + 1:
        raise ValueError("No hay filas de datos entre 'Date' y 'Somme'.")

    return fila_header, fila_somme, col_ultima

def filas_datos(ws, fila_header, fila_somme, col_ultima):
    for r in range(fila_header + 1, fila_somme):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, col_ultima + 1)]
        if any((v is not None and str(v).strip() != "") for v in row_vals):
            yield row_vals

def primera_fila_vacia(ws):
    r = 1
    while True:
        if all((ws.cell(row=r, column=c).value in (None, "")) for c in range(1, ws.max_column + 1)):
            return r
        r += 1

def encontrar_code_lot(ws):
    """
    Busca una celda cuyo valor sea EXACTAMENTE 'Code Lot' y devuelve
    el valor de la celda a su derecha. Si no lo halla, devuelve None.
    """
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if (cell.value is not None) and str(cell.value).strip() == CODE_LOT_LABEL:
                return ws.cell(row=cell.row, column=cell.column + 1).value
    return None

def obtener_o_crear_columna_code_lot(ws_dest, col_ultima_origen=None, creando_headers=False):
    """
    Devuelve el índice de columna en destino para 'Code Lot'.
    - Si ya existe en la fila 1, devuelve su índice.
    - Si no existe, lo crea al final de los encabezados actuales y devuelve su nueva posición.
    - Si el destino está vacío y vamos a crear encabezados, usaremos col_ultima_origen + 1.
    """
    # ¿Destino vacío?
    dest_vacia = all((ws_dest.cell(row=1, column=c).value in (None, "")) for c in range(1, ws_dest.max_column + 1))

    if dest_vacia and creando_headers:
        # La escribiremos al final del header que vamos a crear
        code_col = (col_ultima_origen or 0) + 1
        ws_dest.cell(row=1, column=code_col, value=CODE_LOT_LABEL)
        return code_col

    # Destino con algún encabezado: localizar última col con texto y si ya existe Code Lot
    last_col_idx = 0
    code_lot_col = None
    for cell in ws_dest[1]:
        val = cell.value
        if (val is not None) and str(val).strip() != "":
            last_col_idx = max(last_col_idx, cell.column)
            if str(val).strip() == CODE_LOT_LABEL:
                code_lot_col = cell.column

    if code_lot_col is not None:
        return code_lot_col

    # No existe: crearlo al final
    code_lot_col = last_col_idx + 1 if last_col_idx else 1
    ws_dest.cell(row=1, column=code_lot_col, value=CODE_LOT_LABEL)
    return code_lot_col

def copiar_tabla_de_recepcion_a_destino(archivo_recepcion, ws_destino):
    wb_rec = load_workbook(archivo_recepcion, data_only=True)
    ws_rec = wb_rec.active

    # 1) Límites de datos
    fila_header, fila_somme, col_ultima_origen = encontrar_filas_limite(ws_rec)
    datos = list(filas_datos(ws_rec, fila_header, fila_somme, col_ultima_origen))

    # 2) Code Lot del archivo de origen
    code_lot_val = encontrar_code_lot(ws_rec)  # puede ser None si no existe
    if code_lot_val is None:
        # No abortamos; escribiremos vacío y reportaremos al final
        pass

    if not datos:
        return 0, 0, (code_lot_val is None)

    # 3) ¿Destino vacío? Si sí, creamos encabezados y añadimos "Code Lot"
    start_row = primera_fila_vacia(ws_destino)
    if start_row == 1:
        # Escribir encabezados de origen
        headers = [ws_rec.cell(row=fila_header, column=c).value for c in range(1, col_ultima_origen + 1)]
        for c, val in enumerate(headers, start=1):
            ws_destino.cell(row=1, column=c, value=val)
        # Crear/obtener columna Code Lot (al final)
        code_lot_col = obtener_o_crear_columna_code_lot(ws_destino, col_ultima_origen=col_ultima_origen, creando_headers=True)
        start_row = 2
    else:
        # Destino ya tenía algo: aseguramos columna Code Lot
        code_lot_col = obtener_o_crear_columna_code_lot(ws_destino)

    # 4) Escribir filas + code lot
    r = start_row
    for row_vals in datos:
        # columnas de datos
        for c, val in enumerate(row_vals, start=1):
            ws_destino.cell(row=r, column=c, value=val)
        # code lot al final
        ws_destino.cell(row=r, column=code_lot_col, value=code_lot_val)
        r += 1

    return len(datos), col_ultima_origen, (code_lot_val is None)

def actualizar_destino_desde_varias_recepciones(archivos_recepcion, archivo_destino):
    try:
        wb_dest = load_workbook(archivo_destino)
        ws_dest = wb_dest.active

        total_filas = 0
        ancho_max = 0
        errores = []
        faltas_code_lot = 0

        for path in archivos_recepcion:
            try:
                filas, cols, sin_code = copiar_tabla_de_recepcion_a_destino(path, ws_dest)
                total_filas += filas
                ancho_max = max(ancho_max, cols)
                if sin_code:
                    faltas_code_lot += 1
            except Exception as e:
                errores.append(f"- {path}: {e}")

        wb_dest.save(archivo_destino)

        if errores and total_filas == 0:
            messagebox.showerror("Error", "No se copiaron datos.\n\n" + "\n".join(errores))
        else:
            msg = f"Copiado terminado.\nFilas añadidas: {total_filas}"
            if ancho_max:
                msg += f"\nColumnas origen consideradas: {ancho_max}"
            if faltas_code_lot:
                msg += f"\nArchivos sin 'Code Lot' detectado: {faltas_code_lot}"
            if errores:
                msg += "\n\nIncidencias:\n" + "\n".join(errores)
            messagebox.showinfo("Éxito", msg)

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error: {e}")

def seleccionar_y_ejecutar():
    archivos = filedialog.askopenfilenames(
        title="Selecciona los archivos de recepción",
        filetypes=[("Archivos Excel", "*.xlsx")]
    )
    if not archivos:
        return
    destino = filedialog.askopenfilename(
        title="Selecciona el archivo de destino",
        filetypes=[("Archivos Excel", "*.xlsx")]
    )
    if not destino:
        return
    actualizar_destino_desde_varias_recepciones(archivos, destino)

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Copiar filas y Code Lot")
    root.geometry("460x200")

    tk.Label(root, text="Copiar filas entre 'Date' (exacto) y 'Somme' + añadir 'Code Lot'").pack(pady=10)
    tk.Button(root, text="Seleccionar archivos y ejecutar", command=seleccionar_y_ejecutar, padx=18, pady=8).pack(pady=20)

    root.mainloop()
